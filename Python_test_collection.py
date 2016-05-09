import openpyxl
import xlrd
import xlsxwriter 
import lxml
from lxml import html
import requests

class CardCollection:

    def __init__(self,collectionFile):
  
    #  create member variable self.workingDir to hold directory to use Could either ask who's running the program
    #  create member variable self.collection to hold filename and set it here----> is this done? priority?
    # is this something worth doing? priority is low 
        self.collectionFile = collectionFile
        self.sourceWorkbook = xlrd.open_workbook(self.collectionFile) 
        
    def getNumSheets(self):
        # delete when member variable self.collection created 
        numb_sheets = self.sourceWorkbook.nsheets
        return numb_sheets

    def getNumRows(self, sheetIndex):
        sheet_dest = self.sourceWorkbook.sheet_by_index(sheetIndex)
        numb_rows = sheet_dest.nrows - 1
        return numb_rows

    def getSheetName(self,sheetIndex):
        sh_names_list= self.sourceWorkbook.sheet_names()
        sh_name = sh_names_list[sheetIndex]
        return sh_name
    
    def getCardName(self,sheetIndex,rowIndex):
        
        my_sheet = self.sourceWorkbook.sheet_by_index(sheetIndex)
        my_cardname = my_sheet.cell(rowIndex,1).value
        return my_cardname
    
    def getAllCardInfo(self,sheetIndex,rowIndex):
       
        cardInfo = []
        cardInfo.append(self.getCardName(sheetIndex,rowIndex))
        my_sheet = self.sourceWorkbook.sheet_by_index(sheetIndex)
        #this  gets 5 sequential items 
        for col in range(2,7):
        	# XXXX find a way to convert number to string 1.0 to 1 --> find out where to add it int()
            cardInfo.append(my_sheet.cell(rowIndex,col).value)
        # gets location
        cardInfo.append(my_sheet.cell(rowIndex,11).value)
        #print( 'the list of cardinfo:  ', cardInfo)
        return cardInfo
      
class WebScraperMTG:
    
    def __init__(self,webpage):
  
        self.webpage = webpage
        self.price_list = []
        self.card_list = []
        self.setInformationLists() 

    def setInformationLists(self):

        page = requests.get(self.webpage)
        #below will pars the contents with lxml
        tree = html.fromstring(page.content)
        #below will select elements we need-return a list of elements
        card_list = tree.xpath('//a[@data-full-image]/text()')

        size = int(len(card_list) / 2)
        self.card_list = card_list[0:size]
        #print("card list ", self.card_list)
        
        price_list = tree.xpath('//td[@class="text-right"]/text()')

        price_list = [x for x in price_list if x != '\n']
        size = int(len(price_list)/2)
        self.price_list = []
        for i in range(0,size,3):
            value = price_list[i]
            value = value[1:-1]
            self.price_list.append(value)
            #print("price list",self.price_list)
            #  self.daily_change_list = []
            #  for j in range(1,size,3):
            # value = price_list[j]
            #  value = value[1:-1]
            # self.daily_change_list.append(value)
            #   newSize = int(len(self.daily_change_list))
            # self.daily_change_list = self.daily_change_list[0:newSize]
            #    print(newSize)
            #  print(self.daily_change_list)     
            # self.weekly_change_list = [] for k in range(2,size,3):     value = price_list[j]     value = value[1:-1
            # self.weekly_change_list.append(value) newSize = int(len(self.weekly_change_list)) self.weekly_change_list =
            #  self.weekly_change_list[0:newSize] print(newSize) print(self.weekly_change_list)

    def getWebInformationList(self,cardName):

        index = 0
        print ('cardName inside getWebInformationList' , cardName)
        card_price = "Not Found"
        cardFound = "FALSE"
        lengthOfCard_list = len(self.card_list)
        while (index < lengthOfCard_list) and (cardFound == "FALSE") :
            if cardName == self.card_list[index] :
                card_price = self.price_list[index]
                print ('the index for the  cardname you are looking for is: ', index)
                cardFound = "TRUE"
            index += 1
     
        print ('card price found for this card: ', card_price)
        return card_price

class CardSummary:

    def __init__(self):
        self.currentrow = 1	
        self.setHeader()

    def setHeader(self):

        header_list = ['Card name','Color','Rarity','Foil','Special','Number','Location','Price','Sheet Name']
        self.open_workbook = openpyxl.load_workbook('final_magic.xlsx')
        self.currentsheet = self.open_workbook.active 
        c = 1
        # item is an actual value of the element inside the list and NOT a pointercol =  - lesson learned from how loop work for lists
        for item in header_list:
            self.currentsheet.cell(row= self.currentrow,column=c).value = item
            c += 1
        self.open_workbook.save('final_magic.xlsx') 
        self.currentrow += 1

    def writeSummaryRow(self,cardInfo ,numRows ):
        c = 1
        self.open_workbook = openpyxl.load_workbook('final_magic.xlsx')
        self.currentsheet = self.open_workbook.active

        for item in cardInfo:
            
            self.currentsheet.cell(row= self.currentrow,column=c).value = item
            c += 1    
            self.open_workbook.save('final_magic.xlsx') 
        self.currentrow += 1
    #def saveCardInfo(self): #needs to be adjusted
       
        #new_workbook = xlsxwriter.Workbook('final_magic.xlsx')
        # add one sheet and assign a name-only once
        #new_sheet = new_workbook.add_worksheet('NewZen')
        
        #i = 0
        #for item in cardInfo:
        #    xwrite = new_sheet.write(1,i,cardInfo[i])
        #    i += 1
        #new_workbook.close()
        #print('Card saved to new excel file')

# Card Collection Test Suite  ---implement the test suite with nose

sheetIndex = 46
rowIndex = 1
oCC = CardCollection('/Users/mity/mypy/MTG_Collection_4_20_16.xlsx')
oWS = WebScraperMTG('http://www.mtggoldfish.com/index/ZEN#paper')
oCS = CardSummary()

numSheets = oCC.getNumSheets()
numRows = oCC.getNumRows(sheetIndex)
sheetName = oCC.getSheetName(sheetIndex)


for rowIndex in range (1,numRows): 
    
    cardInfo = oCC.getAllCardInfo(sheetIndex,rowIndex)
    cardName = oCC.getCardName(sheetIndex,rowIndex)

    online_price = oWS.getWebInformationList(cardName)
 
    cardInfo.append(online_price)
    cardInfo.append(sheetName)
    oCS.writeSummaryRow(cardInfo,numRows)


#print('Total number of SHEETS in this file is: ', numSheets)
#print('Total number of ROWS in this sheet is: ',numRows)
print('The SHEET NAME found is: ',sheetName)
print('The CARD NAME is:',cardName)
print('The list of all required information is: ', cardInfo)
print("card price" , online_price)




# if numSheets == 67:
#     print("numSheets Check = TRUE")
# else:
#     ##print("numSheets Check = FALSE")
#     print("numsheet reported is FALSE, expected value is" , numSheets)
    
# if numRows == 323:
#     print("numRows Check = TRUE")
# else:
#     print("numRows Check = FALSE")

# if sheetName == "ZEN":
#     print("sheetName Check = TRUE")
# else:
#     print("sheetName Check = FALSE")

# if cardName == "Bala Ged Thief":
#     print("cardName Check = TRUE")
# else:
#     print("cardName Check = FALSE")

# if cardInfo[0] == 'Bala Ged Thief':
#     print("Cardname Check from list = TRUE")
# else:
#     print("Cardname Check from list = FALSE")

# if cardInfo[1] == 'B':
#      print("Color Check from list = TRUE")
# else:
#      print("Color Check from list = FALSE") 

# if cardInfo[2] == 'Rare':
#     print("Rarity Check from list= TRUE")
# else:
#     print("Rarity Check from list=  FALSE")

# if cardInfo[3] == 'N':   
#     print("Foil Check from list = TRUE")
# else:
#     print("Foil Check from list = FALSE")

# if cardInfo[4] == 'N':   
#     print("Special Check from list = TRUE")
# else:
#     print("Special Check from list = FALSE")

# if cardInfo[5] == '1':  
#     print("Number Check from list = TRUE")
# else:
#     print("Number from list = FALSE")

# if cardInfo[6] == 'Zendikar - Storage Box':   
#     print("Location Check from list = TRUE")
# else:
#     print("Location Check from list = FALSE")
# cardName = oCC.getCardname(47,5)
# cardInfo = oCC.getAllCardInfo(47,5)

# print(cardName)
# if cardName == "Bloodchief Ascension":
#     print("cardname Check = TRUE")
# else:
#     print("cardname Check = FALSE")
